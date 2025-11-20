#!/usr/bin/env python3
"""
Minimal PHI Redaction Engine for Excel Files
Uses Microsoft Presidio to detect and anonymize PII/PHI in Excel workbooks.
"""

import os
import csv
import logging
import hashlib
from typing import Dict, List, Any, Optional
from datetime import datetime
import yaml
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from presidio_analyzer import AnalyzerEngine
from presidio_anonymizer import AnonymizerEngine
from presidio_anonymizer.entities import OperatorConfig


class RedactionEngine:
    """Core redaction engine for Excel files."""

    PHI_COLUMN_INDICATORS = [
        'name', 'patient', 'firstname', 'lastname', 'middlename',
        'dob', 'dateofbirth', 'birthdate', 'birth',
        'address', 'street', 'city', 'state', 'zip', 'zipcode', 'postal',
        'phone', 'telephone', 'mobile', 'cell', 'fax',
        'email', 'mail',
        'ssn', 'social', 'socialsecurity',
        'mrn', 'medicalrecord', 'recordnumber', 'patientid',
        'insurance', 'policy', 'member', 'group', 'subscriber'
    ]

    def __init__(self, config_path: str = None):
        """Initialize engine with configuration."""
        if config_path is None:
            # Default to config/config.yaml relative to project root
            project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            config_path = os.path.join(project_root, "config", "config.yaml")
        self.config = self._load_config(config_path)
        self.analyzer = None
        self.anonymizer = AnonymizerEngine()
        self.detection_log = []
        self._init_analyzer()

        # Set up logging with error handling for file creation
        log_handlers = [logging.StreamHandler()]
        try:
            # Try to create log file in the project directory
            log_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            log_file = os.path.join(log_dir, 'redaction.log')
            log_handlers.append(logging.FileHandler(log_file))
        except Exception as e:
            # If we can't create the log file, just use console output
            print(f"Warning: Could not create log file: {e}")

        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=log_handlers
        )

    def _load_config(self, config_path: str) -> Dict[str, Any]:
        """Load configuration from YAML file."""
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                return yaml.safe_load(f)
        else:
            return self._get_default_config()

    def _get_default_config(self) -> Dict[str, Any]:
        """Return default configuration if no file exists."""
        return {
            'enabled_entities': [
                'PERSON', 'EMAIL_ADDRESS', 'PHONE_NUMBER',
                'US_SSN', 'DATE_TIME', 'LOCATION', 'MEDICAL_LICENSE',
                'NRP', 'CREDIT_CARD', 'IBAN_CODE', 'IP_ADDRESS'
            ],
            'anonymization_strategy': 'replace',  # 'replace' or 'hash'
            'column_redaction_hints': self.PHI_COLUMN_INDICATORS,
            'output_suffix': '_redacted',
            'spacy_model': 'en_core_web_md',
            'confidence_threshold': 0.20,  # Lowered for better name detection
            'custom_recognizers': {
                'enabled': True,
                'mrn_pattern': r'\b[A-Z]{2}\d{6}\b'  # Example: AB123456
            }
        }

    def _init_analyzer(self):
        """Initialize Presidio analyzer with configuration."""
        try:
            # Handle import for both script and module execution
            if __package__ is None or __package__ == '':
                import sys
                project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
                if project_root not in sys.path:
                    sys.path.insert(0, project_root)
                from src.recognizers.custom_recognizers import get_custom_recognizers
            else:
                from ..recognizers.custom_recognizers import get_custom_recognizers

            # Try to initialize the analyzer (this will load the spaCy model)
            self.analyzer = AnalyzerEngine()

            if self.config.get('custom_recognizers', {}).get('enabled', False):
                custom_recognizers = get_custom_recognizers(self.config)
                for recognizer in custom_recognizers:
                    self.analyzer.registry.add_recognizer(recognizer)
        except Exception as e:
            error_msg = str(e)
            if 'en_core_web' in error_msg or 'spacy' in error_msg.lower():
                raise RuntimeError(
                    f"Failed to load spaCy model: {error_msg}\n"
                    "Please install the model by running:\n"
                    "  python -m spacy download en_core_web_md"
                )
            else:
                raise RuntimeError(f"Failed to initialize analyzer: {error_msg}")

    def update_config(self, new_config: Dict[str, Any]):
        """Update configuration at runtime."""
        self.config = new_config
        self._init_analyzer()

    def _should_redact_column(self, header: str) -> bool:
        """Check if column header suggests PHI content."""
        if not header:
            return False

        header_lower = str(header).lower().strip()
        header_clean = ''.join(c for c in header_lower if c.isalnum())

        for indicator in self.config.get('column_redaction_hints', self.PHI_COLUMN_INDICATORS):
            if indicator in header_clean:
                return True

        return False

    def _merge_overlapping_entities(self, results):
        """Merge overlapping entities to avoid partial redaction."""
        if not results:
            return results

        # Sort by start position
        sorted_results = sorted(results, key=lambda x: (x.start, -x.end))

        merged = []
        current = sorted_results[0]

        for next_result in sorted_results[1:]:
            # Check if entities overlap or are adjacent
            if next_result.start <= current.end:
                # Merge: extend the current entity to cover both
                if next_result.end > current.end:
                    # Create a new merged entity with the wider span
                    from presidio_analyzer import RecognizerResult
                    current = RecognizerResult(
                        entity_type=current.entity_type,  # Keep the first entity type
                        start=current.start,
                        end=max(current.end, next_result.end),
                        score=max(current.score, next_result.score)  # Keep higher score
                    )
            else:
                # No overlap, save current and move to next
                merged.append(current)
                current = next_result

        # Don't forget the last entity
        merged.append(current)

        return merged

    def _anonymize_text(self, text: str, entity_type: str = "GENERIC") -> str:
        """Anonymize text based on configured strategy."""
        strategy = self.config.get('anonymization_strategy', 'replace')

        if strategy == 'hash':
            hash_obj = hashlib.md5(text.encode())
            return f"<{entity_type}_{hash_obj.hexdigest()[:8]}>"
        else:  # replace
            return f"<{entity_type}>"

    def analyze_text(self, text: str) -> str:
        """
        Analyze and redact text string.
        Applies all configured recognizers and filters.
        """
        if not text or not str(text).strip():
            return text

        text_str = str(text)
        threshold = self.config.get('confidence_threshold', 0.20)

        results = self.analyzer.analyze(
            text=text_str,
            entities=self.config.get('enabled_entities', []),
            language='en',
            score_threshold=threshold
        )

        if results:
            # Filter out entities that look like JSON artifacts or code
            valid_results = []
            for result in results:
                entity_text = text_str[result.start:result.end]
                if any(char in entity_text for char in ['"', '{', '}', '[', ']', ':']):
                    continue
                valid_results.append(result)

            # Merge overlapping entities
            merged_results = self._merge_overlapping_entities(valid_results)

            redacted = text_str
            for result in sorted(merged_results, key=lambda x: x.start, reverse=True):
                # Log detection (optional, might need context if used outside cell loop)
                # For now, we won't log here to avoid side effects when testing, 
                # or we can log with dummy values if needed.
                # But _analyze_cell handles logging. 
                # Let's return the redacted text and the results so the caller can log.
                
                replacement = self._anonymize_text(
                    text_str[result.start:result.end],
                    result.entity_type
                )
                redacted = redacted[:result.start] + replacement + redacted[result.end:]
            
            return redacted, merged_results
            
        return text_str, []

    def _analyze_cell(self, text: str, row_idx: int, col_idx: int, sheet_name: str) -> Optional[str]:
        """Analyze and redact a single cell."""
        redacted_text, results = self.analyze_text(text)
        
        if results:
            for result in results:
                self.detection_log.append({
                    'sheet': sheet_name,
                    'row': row_idx,
                    'column': col_idx,
                    'entity_type': result.entity_type,
                    'confidence': result.score,
                    'original_length': result.end - result.start
                })
        
        return redacted_text

    def redact_workbook(self, input_path: str, output_path: Optional[str] = None) -> tuple:
        """
        Redact PHI from Excel workbook.
        If a redacted file already exists, it will be replaced.
        Returns: (output_path, detection_report_path)
        """
        logging.info(f"Starting redaction for: {input_path}")
        self.detection_log = []

        if not output_path:
            base_name = os.path.splitext(input_path)[0]
            suffix = self.config.get('output_suffix', '_redacted')
            output_path = f"{base_name}{suffix}.xlsx"

        # Check if redacted file already exists
        if os.path.exists(output_path):
            logging.info(f"Existing redacted file found at {output_path}. It will be replaced.")
            # Create backup of existing file (optional, for safety)
            backup_path = f"{output_path}.backup"
            try:
                if os.path.exists(backup_path):
                    os.remove(backup_path)
                os.rename(output_path, backup_path)
                logging.info(f"Created backup of existing file at {backup_path}")
            except Exception as e:
                logging.warning(f"Could not create backup: {e}")

        workbook = openpyxl.load_workbook(input_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            self._redact_sheet(sheet, sheet_name)

        # Save the new redacted file
        try:
            workbook.save(output_path)
            # If save successful, remove backup
            backup_path = f"{output_path}.backup"
            if os.path.exists(backup_path):
                try:
                    os.remove(backup_path)
                    logging.info("Removed backup file after successful save")
                except:
                    pass  # Ignore errors removing backup
        except Exception as e:
            # If save fails, restore backup
            backup_path = f"{output_path}.backup"
            if os.path.exists(backup_path):
                try:
                    os.rename(backup_path, output_path)
                    logging.error(f"Save failed, restored backup. Error: {e}")
                except:
                    pass
            raise e

        report_path = self._save_detection_report(input_path, output_path)

        logging.info(f"Redaction completed. Output: {output_path}")
        logging.info(f"Detection report: {report_path}")

        return output_path, report_path

    def _redact_sheet(self, sheet: Worksheet, sheet_name: str):
        """Redact PHI from a single sheet."""
        if sheet.max_row == 0:
            return

        headers = []
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            headers.append(header)

        columns_to_redact = []
        for idx, header in enumerate(headers, 1):
            if self._should_redact_column(header):
                columns_to_redact.append(idx)
                logging.info(f"Sheet '{sheet_name}': Column {idx} '{header}' marked for full redaction")

        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)

                if col in columns_to_redact and row > 1:
                    if cell.value:
                        original_type = type(cell.value).__name__
                        self.detection_log.append({
                            'sheet': sheet_name,
                            'row': row,
                            'column': col,
                            'entity_type': 'COLUMN_PHI',
                            'confidence': 1.0,
                            'original_length': len(str(cell.value))
                        })
                        cell.value = self._anonymize_text(str(cell.value), 'COLUMN_PHI')
                else:
                    if isinstance(cell.value, str):
                        redacted = self._analyze_cell(cell.value, row, col, sheet_name)
                        if redacted != cell.value:
                            cell.value = redacted

    def _save_detection_report(self, input_path: str, output_path: str) -> str:
        """Save detection report as CSV. Replaces existing report if present."""
        report_path = output_path.replace('.xlsx', '_report.csv')

        # Check if report already exists
        if os.path.exists(report_path):
            logging.info(f"Existing report found at {report_path}. It will be replaced.")
            try:
                # Create backup of existing report
                backup_path = f"{report_path}.backup"
                if os.path.exists(backup_path):
                    os.remove(backup_path)
                os.rename(report_path, backup_path)
            except Exception as e:
                logging.warning(f"Could not backup existing report: {e}")

        try:
            with open(report_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow([
                    'Timestamp', 'Input_File', 'Output_File', 'Sheet',
                    'Row', 'Column', 'Entity_Type', 'Confidence'
                ])

                timestamp = datetime.now().isoformat()
                for detection in self.detection_log:
                    writer.writerow([
                        timestamp,
                        os.path.basename(input_path),
                        os.path.basename(output_path),
                        detection['sheet'],
                        detection['row'],
                        detection['column'],
                        detection['entity_type'],
                        f"{detection['confidence']:.2f}"
                    ])

            # If save successful, remove backup
            backup_path = f"{report_path}.backup"
            if os.path.exists(backup_path):
                try:
                    os.remove(backup_path)
                    logging.info("Removed report backup after successful save")
                except:
                    pass  # Ignore errors removing backup

        except Exception as e:
            # If save fails, restore backup
            backup_path = f"{report_path}.backup"
            if os.path.exists(backup_path):
                try:
                    os.rename(backup_path, report_path)
                    logging.error(f"Report save failed, restored backup. Error: {e}")
                except:
                    pass
            raise e

        return report_path


def main():
    """CLI interface for redaction engine."""
    import argparse

    parser = argparse.ArgumentParser(description='Redact PHI from Excel files')
    parser.add_argument('input_file', help='Path to input Excel file')
    parser.add_argument('-o', '--output', help='Output file path')
    parser.add_argument('-c', '--config', default=None, help='Config file path (defaults to config/config.yaml)')

    args = parser.parse_args()

    engine = RedactionEngine(config_path=args.config)
    output_path, report_path = engine.redact_workbook(args.input_file, args.output)

    print(f"✓ Redacted file: {output_path}")
    print(f"✓ Report: {report_path}")


if __name__ == '__main__':
    main()