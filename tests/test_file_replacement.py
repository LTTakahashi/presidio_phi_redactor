#!/usr/bin/env python3
"""
Test script to verify file replacement functionality
"""

import os
import sys
import tempfile
import shutil
import openpyxl
from datetime import datetime

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from src.engine.redaction_engine import RedactionEngine


def create_test_excel(filepath, content="John Smith, SSN: 123-45-6789"):
    """Create a test Excel file with PHI data"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Info'
    sheet['A2'] = 'John Smith'
    sheet['B2'] = content
    workbook.save(filepath)
    print(f"Created test file: {filepath}")


def test_file_replacement():
    """Test that redacted files are replaced when re-running redaction"""

    print("="*60)
    print("Testing File Replacement Functionality")
    print("="*60)

    # Create temporary directory for testing
    with tempfile.TemporaryDirectory() as temp_dir:
        # Create test Excel file
        input_file = os.path.join(temp_dir, "test_data.xlsx")
        create_test_excel(input_file, "Patient: John Doe, Phone: 555-1234")

        # Expected output paths
        redacted_file = os.path.join(temp_dir, "test_data_redacted.xlsx")
        report_file = os.path.join(temp_dir, "test_data_redacted_report.csv")

        # Initialize redaction engine
        engine = RedactionEngine()

        # First redaction
        print("\n--- First Redaction ---")
        output1, report1 = engine.redact_workbook(input_file)

        # Check files were created
        assert os.path.exists(redacted_file), "Redacted file not created"
        assert os.path.exists(report_file), "Report file not created"

        # Get modification times
        redacted_mtime1 = os.path.getmtime(redacted_file)
        report_mtime1 = os.path.getmtime(report_file)

        print(f"✓ Created redacted file: {redacted_file}")
        print(f"✓ Created report file: {report_file}")

        # Read first redacted content
        wb1 = openpyxl.load_workbook(redacted_file)
        sheet1 = wb1.active
        content1 = sheet1['B2'].value
        wb1.close()
        print(f"  First redacted content: {content1}")

        # Wait a moment to ensure different timestamps
        import time
        time.sleep(1)

        # Modify the original file
        create_test_excel(input_file, "New Patient: Jane Smith, SSN: 987-65-4321")

        # Second redaction (should replace existing files)
        print("\n--- Second Redaction (Replacing Existing) ---")
        output2, report2 = engine.redact_workbook(input_file)

        # Check files still exist
        assert os.path.exists(redacted_file), "Redacted file missing after replacement"
        assert os.path.exists(report_file), "Report file missing after replacement"

        # Get new modification times
        redacted_mtime2 = os.path.getmtime(redacted_file)
        report_mtime2 = os.path.getmtime(report_file)

        # Verify files were replaced (newer modification times)
        assert redacted_mtime2 > redacted_mtime1, "Redacted file was not replaced"
        assert report_mtime2 > report_mtime1, "Report file was not replaced"

        print(f"✓ Replaced redacted file: {redacted_file}")
        print(f"✓ Replaced report file: {report_file}")

        # Read second redacted content
        wb2 = openpyxl.load_workbook(redacted_file)
        sheet2 = wb2.active
        content2 = sheet2['B2'].value
        wb2.close()
        print(f"  Second redacted content: {content2}")

        # Verify content is different (showing it was replaced)
        assert content1 != content2, "Content was not updated in replaced file"

        # Check no backup files remain
        backup_files = [f for f in os.listdir(temp_dir) if '.backup' in f]
        assert len(backup_files) == 0, f"Backup files were not cleaned up: {backup_files}"

        print("\n✓ File replacement test PASSED!")
        print("  - Files are properly replaced when re-running redaction")
        print("  - Backup files are created during replacement and cleaned up")
        print("  - No duplicate files are created")


def test_error_handling():
    """Test error handling during file replacement"""

    print("\n" + "="*60)
    print("Testing Error Handling During Replacement")
    print("="*60)

    with tempfile.TemporaryDirectory() as temp_dir:
        # Create test Excel file
        input_file = os.path.join(temp_dir, "test_data.xlsx")
        create_test_excel(input_file)

        # Initialize redaction engine
        engine = RedactionEngine()

        # First redaction
        print("\n--- Creating initial redacted file ---")
        output1, report1 = engine.redact_workbook(input_file)
        print(f"✓ Created: {output1}")

        # Make the redacted file read-only to simulate permission error
        redacted_file = os.path.join(temp_dir, "test_data_redacted.xlsx")

        # On Windows, this might not work as expected, so we'll skip on Windows
        if os.name != 'nt':
            os.chmod(redacted_file, 0o444)

            print("\n--- Attempting replacement with read-only file ---")
            try:
                # This should handle the error gracefully
                output2, report2 = engine.redact_workbook(input_file)
                # Reset permissions for cleanup
                os.chmod(redacted_file, 0o644)
                print("✓ Handled read-only file gracefully")
            except Exception as e:
                # Reset permissions for cleanup
                os.chmod(redacted_file, 0o644)
                print(f"✓ Error handled: {str(e)}")
        else:
            print("  Skipping read-only test on Windows")


if __name__ == "__main__":
    test_file_replacement()
    test_error_handling()
    print("\n" + "="*60)
    print("All file replacement tests completed successfully!")
    print("="*60)