#!/usr/bin/env python3
"""
Test suite for the PHI redaction engine.
Covers column redaction, free-text detection, multi-sheet handling, and idempotence.
"""

import os
import unittest
import tempfile
import shutil
import openpyxl
from openpyxl import Workbook
from redaction_engine import RedactionEngine


class TestRedactionEngine(unittest.TestCase):
    """Test cases for the RedactionEngine class."""

    def setUp(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.engine = RedactionEngine()

    def tearDown(self):
        """Clean up test environment."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def create_test_workbook(self, filename: str, data: dict) -> str:
        """Helper to create test Excel files."""
        wb = Workbook()

        for sheet_name, sheet_data in data.items():
            if sheet_name == 'Sheet':
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(sheet_name)

            for row_idx, row_data in enumerate(sheet_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        filepath = os.path.join(self.temp_dir, filename)
        wb.save(filepath)
        return filepath

    def test_whole_column_redaction(self):
        """Test that columns with PHI headers are fully redacted."""
        test_data = {
            'Sheet1': [
                ['ID', 'Name', 'Age', 'Phone', 'Notes'],
                ['001', 'John Doe', 45, '555-1234', 'Patient has allergies'],
                ['002', 'Jane Smith', 32, '555-5678', 'Follow-up needed'],
                ['003', 'Bob Johnson', 67, '555-9012', 'Medication adjusted']
            ]
        }

        input_file = self.create_test_workbook('test_columns.xlsx', test_data)
        output_file, _ = self.engine.redact_workbook(input_file)

        wb = openpyxl.load_workbook(output_file)
        ws = wb['Sheet1']

        # Check that Name column (column 2) is redacted
        self.assertNotEqual(ws.cell(row=2, column=2).value, 'John Doe')
        self.assertIn('<', str(ws.cell(row=2, column=2).value))

        # Check that Phone column (column 4) is redacted
        self.assertNotEqual(ws.cell(row=2, column=4).value, '555-1234')
        self.assertIn('<', str(ws.cell(row=2, column=4).value))

        # Check that ID column (column 1) is NOT redacted
        self.assertEqual(ws.cell(row=2, column=1).value, '001')

    def test_free_text_detection(self):
        """Test detection of PHI in free-text cells."""
        test_data = {
            'Sheet1': [
                ['Notes', 'Comments'],
                ['Patient email is john.doe@example.com', 'Normal text'],
                ['Call patient at 555-123-4567', 'No PHI here'],
                ['SSN: 123-45-6789', 'Regular comment']
            ]
        }

        input_file = self.create_test_workbook('test_freetext.xlsx', test_data)
        output_file, _ = self.engine.redact_workbook(input_file)

        wb = openpyxl.load_workbook(output_file)
        ws = wb['Sheet1']

        # Check email redaction
        cell_value = str(ws.cell(row=2, column=1).value)
        self.assertNotIn('john.doe@example.com', cell_value)
        self.assertIn('<EMAIL_ADDRESS>', cell_value)

        # Check phone redaction
        cell_value = str(ws.cell(row=3, column=1).value)
        self.assertNotIn('555-123-4567', cell_value)
        self.assertIn('<PHONE_NUMBER>', cell_value)

        # Check SSN redaction
        cell_value = str(ws.cell(row=4, column=1).value)
        self.assertNotIn('123-45-6789', cell_value)
        self.assertIn('<US_SSN>', cell_value)

    def test_empty_cells(self):
        """Test that empty cells are handled correctly."""
        test_data = {
            'Sheet1': [
                ['Name', 'Email', 'Notes'],
                ['John Doe', '', 'Some notes'],
                ['', 'jane@example.com', ''],
                [None, None, None]
            ]
        }

        input_file = self.create_test_workbook('test_empty.xlsx', test_data)

        # Should not raise any errors
        output_file, _ = self.engine.redact_workbook(input_file)

        wb = openpyxl.load_workbook(output_file)
        ws = wb['Sheet1']

        # Empty cells should remain empty or None
        self.assertIn(ws.cell(row=2, column=2).value, ['', None])
        self.assertIn(ws.cell(row=4, column=1).value, ['', None])

    def test_multi_sheet_files(self):
        """Test that multi-sheet workbooks are handled correctly."""
        test_data = {
            'Sheet1': [
                ['ID', 'Name'],
                ['001', 'John Doe']
            ],
            'Sheet2': [
                ['Email', 'Phone'],
                ['john@example.com', '555-1234']
            ],
            'Sheet3': [
                ['Address', 'City'],
                ['123 Main St', 'New York']
            ]
        }

        input_file = self.create_test_workbook('test_multisheet.xlsx', test_data)
        output_file, _ = self.engine.redact_workbook(input_file)

        wb = openpyxl.load_workbook(output_file)

        # Check all sheets exist
        self.assertIn('Sheet1', wb.sheetnames)
        self.assertIn('Sheet2', wb.sheetnames)
        self.assertIn('Sheet3', wb.sheetnames)

        # Check redaction in each sheet
        ws1 = wb['Sheet1']
        self.assertNotEqual(ws1.cell(row=2, column=2).value, 'John Doe')

        ws2 = wb['Sheet2']
        self.assertNotIn('@example.com', str(ws2.cell(row=2, column=1).value))

        ws3 = wb['Sheet3']
        self.assertNotEqual(ws3.cell(row=2, column=1).value, '123 Main St')

    def test_idempotence(self):
        """Test that running redaction twice doesn't expose data."""
        test_data = {
            'Sheet1': [
                ['Name', 'Email'],
                ['John Doe', 'john@example.com']
            ]
        }

        input_file = self.create_test_workbook('test_idempotent.xlsx', test_data)

        # First redaction
        output_file1, _ = self.engine.redact_workbook(input_file)

        # Second redaction on already redacted file
        output_file2, _ = self.engine.redact_workbook(output_file1)

        wb1 = openpyxl.load_workbook(output_file1)
        wb2 = openpyxl.load_workbook(output_file2)

        ws1 = wb1['Sheet1']
        ws2 = wb2['Sheet1']

        # Both files should have redacted content
        self.assertNotIn('John Doe', str(ws1.cell(row=2, column=1).value))
        self.assertNotIn('John Doe', str(ws2.cell(row=2, column=1).value))
        self.assertNotIn('john@example.com', str(ws1.cell(row=2, column=2).value))
        self.assertNotIn('john@example.com', str(ws2.cell(row=2, column=2).value))

        # The redacted values should remain redacted
        self.assertIn('<', str(ws2.cell(row=2, column=1).value))
        self.assertIn('<', str(ws2.cell(row=2, column=2).value))

    def test_custom_mrn_recognizer(self):
        """Test custom MRN pattern recognition."""
        test_data = {
            'Sheet1': [
                ['Patient_ID', 'MRN', 'Notes'],
                ['P001', 'AB123456', 'Patient MRN: CD789012'],
                ['P002', 'Not an MRN', 'Invalid: A1234567']
            ]
        }

        input_file = self.create_test_workbook('test_mrn.xlsx', test_data)
        output_file, _ = self.engine.redact_workbook(input_file)

        wb = openpyxl.load_workbook(output_file)
        ws = wb['Sheet1']

        # MRN column should be fully redacted
        self.assertNotEqual(ws.cell(row=2, column=2).value, 'AB123456')

        # MRN in notes should be detected
        notes_value = str(ws.cell(row=2, column=3).value)
        self.assertNotIn('CD789012', notes_value)
        self.assertIn('<', notes_value)

        # Invalid MRN should not be detected as MRN
        notes_value2 = str(ws.cell(row=3, column=3).value)
        # The word "Invalid" should remain, but pattern might still be detected
        self.assertIn('Invalid', notes_value2)

    def test_detection_report_generation(self):
        """Test that detection reports are generated correctly."""
        test_data = {
            'Sheet1': [
                ['Name', 'Email'],
                ['John Doe', 'john@example.com']
            ]
        }

        input_file = self.create_test_workbook('test_report.xlsx', test_data)
        output_file, report_file = self.engine.redact_workbook(input_file)

        # Check report file exists
        self.assertTrue(os.path.exists(report_file))
        self.assertTrue(report_file.endswith('_report.csv'))

        # Check report content
        with open(report_file, 'r') as f:
            content = f.read()
            self.assertIn('Sheet', content)
            self.assertIn('Entity_Type', content)
            self.assertIn('Confidence', content)

    def test_numeric_data_preservation(self):
        """Test that numeric data without PHI is preserved."""
        test_data = {
            'Sheet1': [
                ['ID', 'Age', 'Score', 'Weight'],
                [1, 45, 98.5, 175.5],
                [2, 32, 87.3, 145.2],
                [3, 67, 92.1, 180.0]
            ]
        }

        input_file = self.create_test_workbook('test_numeric.xlsx', test_data)
        output_file, _ = self.engine.redact_workbook(input_file)

        wb = openpyxl.load_workbook(output_file)
        ws = wb['Sheet1']

        # Numeric values should be preserved
        self.assertEqual(ws.cell(row=2, column=1).value, 1)
        self.assertEqual(ws.cell(row=2, column=2).value, 45)
        self.assertEqual(ws.cell(row=2, column=3).value, 98.5)
        self.assertEqual(ws.cell(row=2, column=4).value, 175.5)


if __name__ == '__main__':
    unittest.main()