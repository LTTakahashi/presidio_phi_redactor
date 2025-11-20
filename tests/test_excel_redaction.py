#!/usr/bin/env python3
"""
Test Excel redaction to understand why names appear partially redacted
"""

import os
import sys
import tempfile
import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from src.engine.redaction_engine import RedactionEngine


def test_excel_name_redaction():
    """Test how names are redacted in Excel files"""

    print("="*80)
    print("EXCEL NAME REDACTION TEST")
    print("="*80)

    # Create temporary directory for testing
    with tempfile.TemporaryDirectory() as temp_dir:
        # Create test Excel file with names
        input_file = os.path.join(temp_dir, "test_names.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Test Data"

        # Add test data with various name formats
        test_data = [
            ["Name", "Email", "Phone", "Notes"],
            ["John Smith", "john.smith@example.com", "555-1234", "Patient arrived at 3pm"],
            ["Maria Garcia", "maria@example.com", "555-5678", "Follow up with Maria tomorrow"],
            ["Robert Johnson", "rj@example.com", "555-9012", "Robert needs medication"],
            ["Dr. James Wilson", "jwilson@hospital.com", "555-3456", "Contact Dr. Wilson immediately"],
            ["Jane Doe", "jane.doe@email.com", "555-7890", "Jane's appointment confirmed"],
        ]

        for row_idx, row_data in enumerate(test_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        workbook.save(input_file)
        print(f"Created test file: {input_file}")

        # Test with different thresholds
        thresholds = [0.2, 0.5]

        for threshold in thresholds:
            print(f"\n{'='*60}")
            print(f"Testing with confidence threshold: {threshold}")
            print(f"{'='*60}")

            # Create engine with specific threshold
            engine = RedactionEngine()
            engine.config['confidence_threshold'] = threshold
            engine.config['custom_recognizers']['enabled'] = True
            engine._init_analyzer()

            # Process the file
            output_file = os.path.join(temp_dir, f"redacted_{threshold}.xlsx")
            engine.redact_workbook(input_file, output_file)

            # Read the redacted file
            redacted_wb = openpyxl.load_workbook(output_file)
            redacted_sheet = redacted_wb.active

            # Compare original and redacted
            print("\nRedaction results:")
            print("-"*60)
            print(f"{'Row':<5} {'Column':<10} {'Original':<25} {'Redacted':<30}")
            print("-"*60)

            for row in range(2, sheet.max_row + 1):  # Skip header
                for col in range(1, sheet.max_column + 1):
                    original = sheet.cell(row=row, column=col).value
                    redacted = redacted_sheet.cell(row=row, column=col).value

                    if original != redacted:
                        col_letter = openpyxl.utils.get_column_letter(col)
                        print(f"{row:<5} {col_letter:<10} {str(original):<25} {str(redacted):<30}")

            redacted_wb.close()

            # Analyze specific cells in detail
            print("\n" + "="*60)
            print("DETAILED ANALYSIS OF NAME CELLS")
            print("="*60)

            for row in range(2, sheet.max_row + 1):
                original_name = sheet.cell(row=row, column=1).value  # Name column
                original_notes = sheet.cell(row=row, column=4).value  # Notes column

                # Analyze the name
                print(f"\nRow {row} - Name: '{original_name}'")
                results = engine.analyzer.analyze(
                    text=original_name,
                    entities=['PERSON'],
                    language='en',
                    score_threshold=threshold
                )

                if results:
                    print("  Entities detected in name:")
                    for r in results:
                        entity_text = original_name[r.start:r.end]
                        print(f"    - '{entity_text}' at positions {r.start}-{r.end} (score: {r.score:.2f})")

                    # Simulate redaction
                    redacted_text = original_name
                    for r in sorted(results, key=lambda x: x.start, reverse=True):
                        redacted_text = redacted_text[:r.start] + f"<{r.entity_type}>" + redacted_text[r.end:]
                    print(f"  Simulated redaction: '{redacted_text}'")
                else:
                    print("  No entities detected in name")

                # Analyze the notes
                print(f"\nRow {row} - Notes: '{original_notes}'")
                results_notes = engine.analyzer.analyze(
                    text=original_notes,
                    entities=['PERSON'],
                    language='en',
                    score_threshold=threshold
                )

                if results_notes:
                    print("  Entities detected in notes:")
                    for r in results_notes:
                        entity_text = original_notes[r.start:r.end]
                        print(f"    - '{entity_text}' at positions {r.start}-{r.end} (score: {r.score:.2f})")
                else:
                    print("  No entities detected in notes")


def test_overlapping_entities():
    """Test what happens when entities overlap"""

    print("\n" + "="*80)
    print("OVERLAPPING ENTITIES TEST")
    print("="*80)

    engine = RedactionEngine()
    engine.config['confidence_threshold'] = 0.2
    engine.config['custom_recognizers']['enabled'] = True
    engine._init_analyzer()

    test_cases = [
        "John Smith",
        "Dr. John Smith",
        "John Q. Smith",
        "Smith, John",
    ]

    for text in test_cases:
        print(f"\nTesting: '{text}'")
        results = engine.analyzer.analyze(
            text=text,
            entities=engine.config.get('enabled_entities', []),
            language='en',
            score_threshold=0.2
        )

        if results:
            print("  All entities detected:")
            for r in results:
                entity_text = text[r.start:r.end]
                print(f"    - {r.entity_type}: '{entity_text}' at {r.start}-{r.end} (score: {r.score:.2f})")

            # Show how redaction would work
            redacted = text
            for r in sorted(results, key=lambda x: x.start, reverse=True):
                redacted = redacted[:r.start] + f"<{r.entity_type}>" + redacted[r.end:]
            print(f"  Final redaction: '{redacted}'")


if __name__ == "__main__":
    test_excel_name_redaction()
    test_overlapping_entities()

    print("\n" + "="*80)
    print("TEST COMPLETE")
    print("="*80)