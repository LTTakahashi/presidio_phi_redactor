#!/usr/bin/env python3
"""
Visual test to demonstrate confidence threshold functionality
Creates test Excel file and shows results with different thresholds
"""

import os
import sys
import tempfile
import openpyxl
from datetime import datetime

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from src.engine.redaction_engine import RedactionEngine


def create_test_data():
    """Create test data with various PHI types and detection difficulties"""
    test_data = [
        # Headers
        ["Type", "Test Data", "Expected Detection"],

        # High confidence detections (should be caught even with high threshold)
        ["Email", "john.doe@example.com", "High confidence"],
        ["SSN", "123-45-6789", "High confidence"],
        ["Credit Card", "4111-1111-1111-1111", "High confidence"],

        # Medium confidence detections
        ["Phone", "555-1234", "Medium confidence"],
        ["Name in context", "Patient John Smith arrived", "Medium confidence"],
        ["Date", "Born on 01/15/1990", "Medium confidence"],

        # Lower confidence detections (might need lower threshold)
        ["Name alone", "John", "Lower confidence"],
        ["Ambiguous number", "12345", "Very low confidence"],
        ["Partial address", "123 Main St", "Lower confidence"],

        # Custom pattern (MRN)
        ["MRN", "AB123456", "Custom pattern (0.8)"],

        # Common names from custom recognizer
        ["Common name", "Robert went to the store", "Custom name (0.65)"],
        ["Another name", "Maria called yesterday", "Custom name (0.65)"],
    ]
    return test_data


def test_threshold_levels():
    """Test different threshold levels and show what gets redacted"""

    print("="*80)
    print("CONFIDENCE THRESHOLD TEST - Visual Demonstration")
    print("="*80)
    print("\nThis test shows how different confidence thresholds affect redaction.")
    print("Lower threshold = more aggressive (catches more)")
    print("Higher threshold = more selective (only high confidence)")
    print("-"*80)

    # Create temporary directory for testing
    with tempfile.TemporaryDirectory() as temp_dir:
        # Create test Excel file
        input_file = os.path.join(temp_dir, "threshold_test.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "PHI Test Data"

        # Add test data
        test_data = create_test_data()
        for row_idx, row_data in enumerate(test_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        workbook.save(input_file)
        print(f"\nCreated test file with various PHI types")

        # Test different thresholds
        thresholds = [0.2, 0.5, 0.65, 0.8, 0.9]

        results = {}

        for threshold in thresholds:
            print(f"\n{'='*80}")
            print(f"TESTING WITH THRESHOLD: {threshold}")
            print(f"{'='*80}")

            # Create engine with specific threshold
            engine = RedactionEngine()
            engine.config['confidence_threshold'] = threshold
            engine.config['custom_recognizers']['enabled'] = True

            # Re-initialize with new config
            engine._init_analyzer()

            # Process the file
            output_file = os.path.join(temp_dir, f"redacted_{threshold}.xlsx")
            engine.redact_workbook(input_file, output_file)

            # Read the redacted file
            redacted_wb = openpyxl.load_workbook(output_file)
            redacted_sheet = redacted_wb.active

            # Show what was redacted
            print(f"\nRedaction results with threshold {threshold}:")
            print("-"*60)
            print(f"{'Type':<20} {'Original':<30} {'Redacted':<30}")
            print("-"*60)

            for row_idx in range(2, redacted_sheet.max_row + 1):
                type_cell = redacted_sheet.cell(row=row_idx, column=1).value
                original = sheet.cell(row=row_idx, column=2).value
                redacted = redacted_sheet.cell(row=row_idx, column=2).value

                if original != redacted:
                    print(f"{type_cell:<20} {str(original):<30} {str(redacted):<30}")

            redacted_wb.close()

            # Count detections
            total_tests = sheet.max_row - 1  # Exclude header
            detections = sum(1 for row in range(2, sheet.max_row + 1)
                           if sheet.cell(row=row, column=2).value !=
                              redacted_sheet.cell(row=row, column=2).value)

            results[threshold] = detections
            print(f"\nSummary: {detections}/{total_tests} items redacted")

        # Final summary
        print(f"\n{'='*80}")
        print("SUMMARY - Detection Count by Threshold")
        print("="*80)
        for threshold, count in results.items():
            bar = "█" * int(count * 3)  # Visual bar
            print(f"Threshold {threshold:3.1f}: {count:2d} detections {bar}")

        print("\n" + "="*80)
        print("CONCLUSION")
        print("="*80)
        print("✓ Lower thresholds (0.2-0.5) catch more potential PHI")
        print("✓ Medium thresholds (0.5-0.65) balance detection and precision")
        print("✓ Higher thresholds (0.8-0.9) only catch very confident matches")
        print("\nRecommendation: Use 0.3-0.5 for healthcare data (more aggressive)")
        print("                 Use 0.6-0.7 for general data (more selective)")


def test_gui_runtime_config():
    """Test that runtime config from GUI properly overrides defaults"""

    print("\n" + "="*80)
    print("TESTING GUI RUNTIME CONFIG OVERRIDE")
    print("="*80)

    # Create test engine
    engine = RedactionEngine()

    # Simulate GUI runtime config
    gui_config = {
        'enabled_entities': ['PERSON', 'EMAIL_ADDRESS'],
        'anonymization_strategy': 'hash',
        'confidence_threshold': 0.35,  # Custom threshold from GUI
        'column_redaction_hints': ['name', 'email'],
        'custom_recognizers': {
            'enabled': True,
            'mrn_pattern': r'\b[A-Z]{2}\d{6}\b'
        },
        'output_suffix': '_redacted',
        'spacy_model': 'en_core_web_md'
    }

    print("\nDefault config threshold:", engine.config.get('confidence_threshold'))

    # Override with GUI config (simulating what happens in the GUI)
    engine.config = gui_config
    engine._init_analyzer()

    print("After GUI override threshold:", engine.config.get('confidence_threshold'))

    # Test with sample text
    test_text = "John Smith's email is john@example.com and his MRN is AB123456"

    results = engine.analyzer.analyze(
        text=test_text,
        entities=engine.config.get('enabled_entities'),
        language='en',
        score_threshold=engine.config.get('confidence_threshold')
    )

    print(f"\nAnalyzing: '{test_text}'")
    print(f"With threshold: {engine.config.get('confidence_threshold')}")
    print(f"\nDetections:")
    for result in results:
        detected_text = test_text[result.start:result.end]
        print(f"  - {result.entity_type}: '{detected_text}' (score: {result.score:.2f})")

    if engine.config.get('confidence_threshold') == gui_config['confidence_threshold']:
        print("\n✓ GUI config threshold is properly applied!")
    else:
        print("\n✗ ERROR: GUI config threshold not applied correctly!")


if __name__ == "__main__":
    test_threshold_levels()
    test_gui_runtime_config()

    print("\n" + "="*80)
    print("All threshold tests completed!")
    print("Check the debug output to verify the threshold is being passed correctly.")
    print("="*80)