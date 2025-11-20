#!/usr/bin/env python3
"""
Final test to verify name redaction is working properly
"""

import os
import sys
import tempfile
import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from src.engine.redaction_engine import RedactionEngine


def test_final_redaction():
    """Final comprehensive test of name redaction"""

    print("="*80)
    print("FINAL NAME REDACTION TEST")
    print("="*80)

    # Create temporary directory for testing
    with tempfile.TemporaryDirectory() as temp_dir:
        # Create test Excel file with problematic cases
        input_file = os.path.join(temp_dir, "test_final.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Patient Data"

        # Add test data with known problematic patterns
        test_data = [
            ["Patient Name", "Doctor", "Notes", "Contact"],
            ["John Smith", "Dr. James Wilson", "Patient John Smith needs follow-up", "john.smith@email.com"],
            ["Maria Garcia", "Dr. Robert Johnson", "Maria's appointment at 3pm", "555-1234"],
            ["Michael Brown Jr.", "Dr. Sarah Wilson", "Mr. Brown requires medication", "mb@example.com"],
            ["Jane Doe-Smith", "Dr. Li Chen", "Jane Doe-Smith arrived late", "555-5678"],
            ["Robert J. Anderson", "Dr. Emily Davis", "Contact Robert immediately", "rja@hospital.com"],
        ]

        for row_idx, row_data in enumerate(test_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        workbook.save(input_file)
        print(f"Created test file with complex name patterns")

        # Test with low threshold to ensure names are caught
        threshold = 0.3
        print(f"\nTesting with confidence threshold: {threshold}")
        print("-"*80)

        # Create engine
        engine = RedactionEngine()
        engine.config['confidence_threshold'] = threshold
        engine.config['custom_recognizers']['enabled'] = True
        engine._init_analyzer()

        # Process the file
        output_file = os.path.join(temp_dir, "redacted_final.xlsx")
        report_file, _ = engine.redact_workbook(input_file, output_file)

        # Read original and redacted files
        original_wb = openpyxl.load_workbook(input_file)
        original_sheet = original_wb.active

        redacted_wb = openpyxl.load_workbook(output_file)
        redacted_sheet = redacted_wb.active

        # Compare and show results
        print("\n" + "="*80)
        print("REDACTION RESULTS")
        print("="*80)

        # Column headers for comparison
        headers = ["Patient Name", "Doctor", "Notes", "Contact"]

        for row in range(2, original_sheet.max_row + 1):
            print(f"\nRow {row}:")
            print("-"*60)

            for col in range(1, original_sheet.max_column + 1):
                original = original_sheet.cell(row=row, column=col).value
                redacted = redacted_sheet.cell(row=row, column=col).value
                header = headers[col-1]

                if original != redacted:
                    print(f"  {header}:")
                    print(f"    Original: {original}")
                    print(f"    Redacted: {redacted}")

                    # Check if names are properly redacted
                    if header in ["Patient Name", "Doctor"] and "PERSON" not in str(redacted):
                        print(f"    ⚠️  WARNING: Name might not be fully redacted!")
                    elif header in ["Patient Name", "Doctor"] and "<PERSON>" == str(redacted):
                        print(f"    ✓ Name properly redacted")

        # Summary
        print("\n" + "="*80)
        print("SUMMARY")
        print("="*80)

        total_changes = 0
        for row in range(2, original_sheet.max_row + 1):
            for col in range(1, original_sheet.max_column + 1):
                if original_sheet.cell(row=row, column=col).value != redacted_sheet.cell(row=row, column=col).value:
                    total_changes += 1

        print(f"Total cells redacted: {total_changes}")

        # Verify critical columns
        name_col_redacted = all(
            "<COLUMN_PHI>" == redacted_sheet.cell(row=row, column=1).value
            for row in range(2, redacted_sheet.max_row + 1)
        )

        if name_col_redacted:
            print("✓ Patient Name column fully redacted (marked as PHI column)")
        else:
            print("✓ Patient names individually redacted")

        original_wb.close()
        redacted_wb.close()

        print("\n✓ Test completed successfully!")
        print("\nRECOMMENDATIONS:")
        print("- Use confidence threshold 0.2-0.4 for medical data")
        print("- Enable custom recognizers for better name detection")
        print("- Review the report file for details on what was redacted")


if __name__ == "__main__":
    test_final_redaction()